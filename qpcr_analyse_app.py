import streamlit as st
import pandas as pd
import numpy as np
import json
import plotly.graph_objects as go
import plotly.io as pio
from plotly.express import colors as px_colors
import io
import base64
import openpyxl
from io import BytesIO
from scipy import stats
# --- Additional imports for report generation ---
from fpdf import FPDF
import base64

from pydrive2.auth import GoogleAuth
from pydrive2.drive import GoogleDrive
import streamlit as st

@st.cache_resource
def init_drive():
    cfg = json.loads(st.secrets["gdrive_oauth"]["client_config"])
    gauth = GoogleAuth()
    gauth.settings["client_config"] = cfg
    gauth.settings["client_config_backend"] = "settings"
    gauth.settings["get_refresh_token"] = True
    gauth.LocalWebserverAuth()
    return GoogleDrive(gauth)

drive = init_drive()

# Задайте ID папки в вашем Google Drive для хранения
FOLDER_ID = "1fs3_8TS2gCvIWg4LBWc445JAyB6xluq6"

# Set Streamlit page config after all imports
st.set_page_config(layout="wide")
# --- Выбор языка / Language selector ---
lang = st.sidebar.selectbox("Language / Язык", ["English", "Русский"], index=0)
# Переводы ключевых строк
translations = {
    "Русский": {
        "title": "🧪 qPioneer qPCR app",
        "markup_header": "💾 Разметка",
        "upload_markup": "Загрузить разметку (JSON)",
        "markup_loaded": "Разметка загружена",
        "markup_load_error": "Ошибка загрузки разметки",
        "save_markup_button": "💾 Сохранить разметку",
        "restore_session_button": "Восстановить сессию",
        "download_markup_json": "Скачать разметку JSON",
        "load_header": "📥 Загрузка данных qPCR",
        "load_prompt": "Выберите qPCR-файл",
        "load_info":   "⬆️ Загрузите файл для начала.",
        "load_error_cols": "Файл должен содержать колонки 'Pos' и 'Cp'",
        "load_success": "⬆️ Успешно загружено {n} лунок.",
        "settings_header": "⚙️ Настройки",
        "run_id_label": "Run ID",
        "plate_type_label": "Тип планшетки",
        "font_size_label": "Размер шрифта в heatmap",
        "tab1": "1️⃣ Гены",
        "tab2": "2️⃣ Шаблон и репликаты",
        "step1_header": "🧩 Шаг 1: Назначение генов",
        "role_label": "Роль",
        "gene_label": "Ген",
        "select_region_genes": "Выбрать область (гены)",
        "add_region_genes": "➕ Добавить область (гены)",
        "step2_header": "🧩 Шаг 2: Назначение шаблона и репликатов",
        "template_label": "Шаблон (condition)",
        "replicate_label": "Номер репликата",
        "select_region_reps": "Выбрать область (репликаты)",
        "add_region_reps": "➕ Добавить область (репликаты)",
        "status": "🔍 Статус разметки планшетки",
        "calc_expression": "📊 Рассчитать экспрессию",
        "results_header": "📋 Результаты экспрессии",
        "download_csv": "📥 Скачать CSV",
        "download_json": "📥 Скачать JSON результатов",
        "download_excel": "📥 Скачать Excel",
        "boxplot_header": "📊 Boxplots экспрессии по условиям и генам",
        "download_boxplot": "📥 Скачать {g}.png",
        "kaleido_warning": "Установите kaleido (`pip install kaleido`) для скачивания.",
        "select_region_from": "От (верхний левый)",
        "select_region_to": "До (нижний правый)",
        "file_format_header": "Формат файла и колонки",
        "delimiter_label": "Разделитель / Delimiter",
        "skiprows_label": "Пропустить строк / Skip header rows",
        "pos_col_label": "Колонка позиции / Well‑position column",
        "cp_col_label": "Колонка Ct‑значения / Ct value column",
        "load_error_cols_map": "Колонки '{pos}' и '{cp}' должны быть в файле",
        "plot_type_label": "Тип графика",
        "barplot_header": "📊 Барплот с ошибками",
        "outlier_threshold_label": "Порог выбросов (σ)",
        "qc_summary_header": "📈 QC‑отчет",
        "qc_total_wells": "Всего лунок",
        "qc_outliers": "Выбросы",
        "tab_help": "Начало работы",
        "help_getting_started": """
**Как начать работу**

1. Загрузите файл qPCR (txt/csv/xlsx).
2. Настройте формат: разделитель, заголовки, колонки.
3. На вкладках “Гены” и “Шаблон и репликаты” разметьте лунки.
4. Задайте порог выбросов и нажмите “📊 Рассчитать экспрессию”.
5. Просмотрите QC-отчет и графики, скачайте результаты.
""",
        "help_delimiter": "Выберите разделитель, используемый в файле.",
        "help_skiprows": "Количество строк заголовка для пропуска.",
        "help_pos_col": "Имя колонки с позицией лунки (например, A1).",
        "help_cp_col": "Имя колонки с Ct-значениями.",
        "help_calc_expression": "Нажмите, чтобы рассчитать выражение (ΔCt).",
        "help_plot_type": "Выберите тип графика: boxplot или barplot с ошибками.",
    },
    "English": {
        "title": "🧪 qPioneer qPCR app",
        "markup_header": "💾 Annotation",
        "upload_markup": "Load annotation (JSON)",
        "markup_loaded": "Annotation loaded",
        "markup_load_error": "Annotation load error",
        "save_markup_button": "💾 Save annotation",
        "restore_session_button": "Restore session",
        "download_markup_json": "Download annotation JSON",
        "load_header": "📥 Upload qPCR Data",
        "load_prompt": "Select the qPCR data file",
        "load_info":   "⬆️ Please upload a file to begin.",
        "load_error_cols": "File must contain 'Pos' and 'Cp' columns",
        "load_success":"⬆️ Successfully loaded {n} wells.",
        "settings_header": "⚙️ Settings",
        "run_id_label": "Run ID",
        "plate_type_label": "Plate type",
        "font_size_label": "Font size in heatmap",
        "tab1": "1️⃣ Genes",
        "tab2": "2️⃣ Condition & Replicates",
        "step1_header": "🧩 Step 1: Assign Genes",
        "role_label": "Role",
        "gene_label": "Gene",
        "select_region_genes": "Select region (genes)",
        "add_region_genes": "➕ Add region (genes)",
        "step2_header": "🧩 Step 2: Assign Condition & Replicates",
        "template_label": "Condition",
        "replicate_label": "Replicate number",
        "select_region_reps": "Select region (replicates)",
        "add_region_reps": "➕ Add region (replicates)",
        "status": "🔍 Annotation Status",
        "calc_expression": "📊 Calculate expression",
        "results_header": "📋 Expression Results",
        "download_csv": "📥 Download CSV",
        "download_json": "📥 Download JSON Results",
        "download_excel": "📥 Download Excel",
        "boxplot_header": "📊 Expression boxplots by Condition & Gene",
        "download_boxplot": "📥 Download {g}.png",
        "kaleido_warning": "Install kaleido (`pip install kaleido`) to enable downloads.",
        "select_region_from": "From (top left)",
        "select_region_to": "To (bottom right)",
        "file_format_header": "File Format & Columns",
        "delimiter_label": "Delimiter / Разделитель",
        "skiprows_label": "Skip header rows / Пропустить строк",
        "pos_col_label": "Well‑position column / Колонка позиции",
        "cp_col_label": "Ct value column / Колонка Ct‑значения",
        "load_error_cols_map": "Columns '{pos}' and '{cp}' must exist in file",
        "load_prompt": "Select the qPCR data file",
        "load_info":   "⬆️ Please upload a file to begin.",
        "load_success":"⬆️ Successfully loaded {n} wells.",
        "load_error_cols_map": "Columns '{pos}' and '{cp}' must exist in file",
        "plot_type_label": "Plot type",
        "barplot_header": "📊 Barplot with error bars",
        "outlier_threshold_label": "Outlier threshold (σ)",
        "qc_summary_header": "📈 QC Summary",
        "qc_total_wells": "Total wells",
        "qc_outliers": "Outliers",
        "tab_help": "Getting Started",
        "help_getting_started": """
**Quickstart Guide**

1. Upload your qPCR data file (txt/csv/xlsx).
2. Configure file format: delimiter, header rows, column names.
3. Annotate wells in “Genes” and “Condition & Replicates” tabs.
4. Set outlier threshold and click “📊 Calculate expression”.
5. Review QC summary and plots, download your results.
""",
        "help_delimiter": "Select the delimiter used in your file.",
        "help_skiprows": "Number of header rows to skip when reading the file.",
        "help_pos_col": "Column name with well positions (e.g., A1).",
        "help_cp_col": "Column name with Ct values.",
        "help_calc_expression": "Click to normalize and compute expression (ΔCt).",
        "help_plot_type": "Choose chart type: boxplot or barplot with error bars.",
    }
}
texts = translations[lang]

# --- Getting Started Expander ---
with st.sidebar.expander(texts["tab_help"], expanded=True):
    st.markdown(texts["help_getting_started"])

# --- Save / Load entire session state ---
st.sidebar.markdown("## 🔄 Session")
# Load session
sess_file = st.sidebar.file_uploader("Load session (JSON)", type=["json"])
if sess_file:
    try:
        sess = json.loads(sess_file.read().decode("utf-8"))
        # restore all persistent keys
        for key in ("annotation", "run", "delimiter", "plate_type", "font_size", "merged_t"):
            if key in sess:
                if key == "merged_t":
                    # Restore merged_t as a pandas DataFrame
                    import pandas as _pd
                    st.session_state["merged_t"] = _pd.DataFrame(sess["merged_t"])
                else:
                    st.session_state[key] = sess[key]
        # Restore uploaded_runs if present
        if "uploaded_runs" in sess:
            st.session_state["uploaded_runs"] = {
                name: base64.b64decode(data.encode("ascii"))
                for name, data in sess["uploaded_runs"].items()
            }
        st.sidebar.success("Session restored")
        # Removed broken restore session rerun logic
        # if st.sidebar.button(texts["restore_session_button"]):
        #     raise RerunException({})
    except Exception as e:
        st.sidebar.error(f"Failed to load session: {e}")

# Save session
if st.sidebar.button("Save session"):
    # prepare session dict
    sess = {
        "annotation": st.session_state.get("annotation", {}),
        "run": st.session_state.get("run", ""),
        "delimiter": st.session_state.get("delimiter", "\t"),
        "skip_rows": st.session_state.get("skip_rows", 1),
        "pos_col": st.session_state.get("pos_col", "Pos"),
        "cp_col": st.session_state.get("cp_col", "Cp"),
        "plate_type": st.session_state.get("plate_type", "384-well"),
        "font_size": st.session_state.get("font_size", 8),
        # merged_t may be DataFrame; convert to records
        "merged_t": st.session_state.get("merged_t", pd.DataFrame()).to_dict(orient="records"),
        "uploaded_runs": {
            name: base64.b64encode(data).decode("ascii")
            for name, data in st.session_state.get("uploaded_runs", {}).items()
        }
    }
    sess_json = json.dumps(sess, ensure_ascii=False).encode("utf-8")
    st.sidebar.download_button(
        "Download session (JSON)",
        data=sess_json,
        file_name="session.json",
        mime="application/json",
    )

st.title(texts["title"])

# --- Разметка: загрузка / сохранение ---
st.sidebar.header(texts["markup_header"])
# Загрузка разметки из JSON
annot_file = st.sidebar.file_uploader(texts["upload_markup"], type=["json"])
if annot_file:
    try:
        data = json.loads(annot_file.read().decode("utf-8"))
        st.session_state.annotation = data
        # Сразу после загрузки JSON: перекладываем старые роли в новые
        role_map = {
            'water': 'negative control',
            'reference': 'reference gene',
            'target': 'target gene'
        }
        for w, ent in st.session_state.annotation.items():
            old = ent.get('role')
            if old in role_map:
                ent['role'] = role_map[old]
            st.session_state.annotation[w] = ent
        st.sidebar.success(texts["markup_loaded"])
    except Exception as e:
        st.sidebar.error(texts["markup_load_error"] + f": {e}")

# Инициализируем, если ещё нет
if "annotation" not in st.session_state:
    st.session_state.annotation = {}

# Кнопка для экспорта текущей разметки
if st.sidebar.button(texts["save_markup_button"]):
    annot_json = json.dumps(st.session_state.annotation, ensure_ascii=False).encode("utf-8")
    st.sidebar.download_button(
        texts["download_markup_json"],
        data=annot_json,
        file_name="annotation.json",
        mime="application/json",
    )


# --- File Format & Columns ---
st.sidebar.header(texts["file_format_header"])
delimiter = st.sidebar.selectbox(
    texts["delimiter_label"],
    options=["\t", ",", ";", " "],
    format_func=lambda x: {"\t":"Tab", ",":"Comma", ";":"Semicolon", " ":"Space"}[x],
    index=0,
    help=texts["help_delimiter"]
)
skip_rows = st.sidebar.number_input(
    texts["skiprows_label"],
    min_value=0, max_value=10,
    value=st.session_state.get("skip_rows", 1),
    step=1,
    key="skip_rows",
    help=texts["help_skiprows"]
)

pos_col = st.sidebar.text_input(
    texts["pos_col_label"],
    value=st.session_state.get("pos_col", "Pos"),
    key="pos_col",
    help=texts["help_pos_col"]
)
cp_col  = st.sidebar.text_input(
    texts["cp_col_label"],
    value=st.session_state.get("cp_col", "Cp"),
    key="cp_col",
    help=texts["help_cp_col"]
)


# -------------------- qPCR Data Load --------------------
st.header(texts["load_header"])
# If we already have uploaded_runs in session, skip re‑upload
if "uploaded_runs" in st.session_state and st.session_state["uploaded_runs"]:
    run_names = list(st.session_state["uploaded_runs"].keys())
    # allow selection among restored runs
    selected_run = st.sidebar.selectbox("Select Run", run_names, index=0, key="run_select")
else:
    # first‑time upload
    uploaded_files = st.file_uploader(
        texts["load_prompt"], 
        type=["txt","csv","xls","xlsx"], 
        accept_multiple_files=True
    )
    if uploaded_files:
        for f in uploaded_files:
            st.session_state.setdefault("uploaded_runs", {})[f.name] = f.getvalue()
    # if still no data, prompt and exit
    if not st.session_state.get("uploaded_runs"):
        st.info(texts["load_info"])
        st.stop()
    run_names = list(st.session_state["uploaded_runs"].keys())
    selected_run = st.sidebar.selectbox("Select Run", run_names, index=0, key="run_select")

# Load the chosen run from memory
raw = st.session_state["uploaded_runs"][selected_run]
uploaded = io.BytesIO(raw)
uploaded.name = selected_run

def load_df(f, sep, skip):
    ext = f.name.split(".")[-1].lower()
    if ext in ("xls", "xlsx"):
        return pd.read_excel(f, skiprows=skip)
    elif ext == "csv":
        return pd.read_csv(f, sep=",", skiprows=skip)
    else:
        return pd.read_csv(f, sep=sep, skiprows=skip)

df = load_df(uploaded, sep=delimiter, skip=skip_rows)
df.columns = df.columns.str.strip()
if not {pos_col, cp_col}.issubset(df.columns):
    st.error(texts["load_error_cols_map"].format(pos=pos_col, cp=cp_col))
    st.stop()
df["Pos"] = df[pos_col]
df["Cp"]  = df[cp_col]

df["Cp"] = df["Cp"].replace(0, np.nan)
df["row"] = df["Pos"].str[0]
df["col"] = df["Pos"].str[1:].astype(int)
df["well"] = df["row"] + df["col"].astype(str)

st.success(texts["load_success"].format(n=len(df)))
st.dataframe(df.head(), use_container_width=True)  # оставляем, не переводится

# --- Инициализация state для разметки ---
if "annotation" not in st.session_state:
    st.session_state.annotation = {}

# Сетка 384-well
rows = list("ABCDEFGHIJKLMNOP")
cols = list(range(1, 25))
wells = [f"{r}{c}" for r in rows for c in cols]

# --- Общие настройки ---
st.sidebar.markdown(f"## {texts['settings_header']}")
st.sidebar.text_input(texts["run_id_label"], key="run", value="Run1")
# -------------------- Тип планшетки --------------------
plate_type = st.sidebar.selectbox(texts["plate_type_label"], ("384-well", "96-well"), index=0)

if plate_type == "384-well":
    rows = list("ABCDEFGHIJKLMNOP")
    cols = list(range(1, 25))
else:
    rows = list("ABCDEFGH")
    cols = list(range(1, 13))

wells = [f"{r}{c}" for r in rows for c in cols]
# ------------------------------------------------------
font_size = st.sidebar.slider(texts["font_size_label"], min_value=6, max_value=20, value=8, step=1)

# Вкладки для двух этапов разметки
tab1, tab2 = st.tabs([texts["tab1"], texts["tab2"]])

# -------------------- Вкладка 1: Гены --------------------
with tab1:
    st.header(texts["step1_header"])
    gene = st.text_input(texts["gene_label"], value="ACTB")
    role = st.selectbox(texts["role_label"], ["reference gene", "target gene", "negative control"])
    st.subheader(texts["select_region_genes"])
    c1, c2 = st.columns(2)
    start1 = c1.selectbox(texts["select_region_from"], wells, key="g_start")
    end1   = c2.selectbox(texts["select_region_to"], wells, key="g_end")
    if st.button(texts["add_region_genes"]):
        r0, c0 = start1[0], int(start1[1:])
        r1, c1n = end1[0],   int(end1[1:])
        i0, i1 = rows.index(r0), rows.index(r1)
        c0, c1n = min(c0, c1n), max(c0, c1n)
        cnt = 0
        for rr in rows[min(i0, i1): max(i0, i1) + 1]:
            for cc in range(c0, c1n + 1):
                w = f"{rr}{cc}"
                ent = st.session_state.annotation.get(w, {})
                ent.update({"gene": gene, "role": role})
                st.session_state.annotation[w] = ent
                cnt += 1
        st.success(f"Привязано {cnt} лунок к гену «{gene}».")

# -------------------- Вкладка 2: Шаблон и репликаты --------------------
with tab2:
    st.header(texts["step2_header"])
    template = st.text_input(texts["template_label"], value="N4E8")
    replicate = st.text_input(texts["replicate_label"], value="1")
    st.subheader(texts["select_region_reps"])
    c3, c4 = st.columns(2)
    start2 = c3.selectbox(texts["select_region_from"], wells, key="s_start")
    end2   = c4.selectbox(texts["select_region_to"], wells, key="s_end")
    if st.button(texts["add_region_reps"]):
        r0, c0 = start2[0], int(start2[1:])
        r1, c1n = end2[0],   int(end2[1:])
        i0, i1 = rows.index(r0), rows.index(r1)
        c0, c1n = min(c0, c1n), max(c0, c1n)
        cnt = 0
        for rr in rows[min(i0, i1): max(i0, i1) + 1]:
            for cc in range(c0, c1n + 1):
                w = f"{rr}{cc}"
                ent = st.session_state.annotation.get(w, {})
                ent.update({
                    "template": template,
                    "replicate": replicate,
                    "run": st.session_state.run
                })
                st.session_state.annotation[w] = ent
                cnt += 1
        st.success(f"Привязано {cnt} лунок к шаблону «{template}», репликат #{replicate}.")

# -------------------- Итоговая heatmap --------------------
st.markdown("---")
st.subheader(texts["status"])

z = []
text = []
palette = {
    "negative control": "blue",
    "reference gene": "green",
    "target gene": "red"
}

for r in rows:
    for c in cols:
        w = f"{r}{c}"
        ent = st.session_state.annotation.get(w, {})
        if ent.get("role"):
            # Универсальный маппинг с учётом старых и новых ролей
            role_key = ent.get('role', '')
            map_to = {
                'negative control': 1,
                'reference gene': 2,
                'target gene': 3
            }
            z.append(map_to.get(role_key, 0))
            text.append(
                f"{ent.get('gene','')}<br>"
                f"{ent.get('template','')}<br>"
                f"rep: {ent.get('replicate','')}"
)
        else:
            z.append(0)
            text.append("")

# «len(rows)×len(cols)» вместо жёстких 16×24
z_mat    = np.array(z).reshape(len(rows), len(cols))[::-1, :]
text_mat = np.array(text).reshape(len(rows), len(cols))[::-1, :]

fig = go.Figure(go.Heatmap(
    z=z_mat,
    text=text_mat,
    hoverinfo="text",
    texttemplate="%{text}",           # показываем текст прямо в ячейках
    textfont=dict(size=font_size),    # динамический размер шрифта
    colorscale=[
        [0, "white"],
        [1/3, palette["negative control"]],
        [2/3, palette["reference gene"]],
        [1, palette["target gene"]],
    ],
    showscale=False
))
fig.update_traces(xgap=1, ygap=1)
fig.update_layout(
    height=600,
    xaxis=dict(
        tickmode="array",
        tickvals=list(range(len(cols))),
        ticktext=list(range(1, len(cols)+1))
    ),
    yaxis=dict(
        tickmode="array",
        tickvals=list(range(len(rows))),
        ticktext=rows[::-1]
    ),
    margin=dict(t=10, b=10, l=10, r=10)
)
st.plotly_chart(fig, use_container_width=True)

#
# -------------------- Расчёт экспрессии и QC --------------------
# Threshold slider for outlier detection
threshold = st.slider(
    texts["outlier_threshold_label"],
    min_value=0.0, max_value=5.0,
    value=2.0, step=0.5
)
if st.button(texts["calc_expression"], help=texts["help_calc_expression"]):
    ann_df = (
        pd.DataFrame.from_dict(st.session_state.annotation, orient="index")
          .reset_index().rename(columns={"index":"well"})
    )
    dfm = df.merge(ann_df, on="well", how="left")
    dfm = dfm[dfm["role"].isin(["reference gene","target gene"])].dropna(subset=["Cp"])
    # Outlier detection
    dfm["tech_mean"] = dfm.groupby(["template","replicate","gene","run"])["Cp"].transform("mean")
    dfm["tech_sd"]   = dfm.groupby(["template","replicate","gene","run"])["Cp"].transform("std")
    dfm["is_outlier"] = (dfm["Cp"] - dfm["tech_mean"]).abs() > threshold * dfm["tech_sd"]
    # QC summary
    total_wells = len(dfm)
    outliers    = int(dfm["is_outlier"].sum())
    pass_rate   = (total_wells - outliers) / total_wells * 100 if total_wells>0 else 0
    outlier_rate = outliers / total_wells * 100 if total_wells > 0 else 0
    st.session_state["qc_summary"] = {
        "total_wells": total_wells,
        "outliers": outliers,
        "pass_rate": pass_rate,
        "outlier_rate": outlier_rate
    }

    # 1) Среднее Cp по техническим репликам
    tech_avg = (
        dfm
        .groupby(["template","replicate","gene","run"])["Cp"]
        .mean()
        .rename("Cp_avg")
        .reset_index()
    )

    # 2) Среднее Cp по референсам
    ref_avg = (
        dfm[dfm["role"] == "reference gene"]
        .groupby(["template","run"])["Cp"]
        .mean()
        .rename("Cp_ref")
        .reset_index()
    )

    # 3) Слияние и расчёт ΔCt / expression
    merged = tech_avg.merge(ref_avg, on=["template","run"], how="left")
    merged["ΔCt"]      = merged["Cp_avg"] - merged["Cp_ref"]
    merged["expression"] = 2 ** (-merged["ΔCt"])

    # Оставляем только target-гены
    merged_t = merged[merged["gene"].isin(
        dfm[dfm["role"]=="target gene"]["gene"].unique()
    )]

    from itertools import combinations

    # Pairwise t-tests per gene, Bonferroni correction
    pair_pvals = {}
    for gene in merged_t["gene"].unique():
        subs = merged_t[merged_t["gene"] == gene]
        templates = subs["template"].unique().tolist()
        comps = list(combinations(templates, 2))
        m = len(comps)
        gene_pairs = []
        for a, b in comps:
            vals_a = subs[subs["template"] == a]["expression"]
            vals_b = subs[subs["template"] == b]["expression"]
            if len(vals_a)>1 and len(vals_b)>1:
                p = stats.ttest_ind(vals_a, vals_b, equal_var=False).pvalue
            else:
                p = None
            # Bonferroni adjust
            if p is not None:
                p_adj = min(p * m, 1.0)
            else:
                p_adj = None
            gene_pairs.append(((a, b), p_adj))
        pair_pvals[gene] = gene_pairs
    st.session_state["pair_pvals"] = pair_pvals
    st.session_state["merged_t"] = merged_t

#
# Persist merged results and allow chart selection
if "merged_t" in st.session_state:
    merged_t = st.session_state["merged_t"]

    # Let user choose plot type
    plot_type = st.radio(
        texts["plot_type_label"],
        [texts["boxplot_header"], texts["barplot_header"]],
        key="plot_type",
        help=texts["help_plot_type"]
    )

    # Always show full results including any p-values
    st.subheader(texts["results_header"])
    st.dataframe(
        merged_t,
        use_container_width=True
    )
    # --- CSV and JSON download buttons ---
    st.download_button(
        texts["download_csv"],
        data=merged_t.to_csv(index=False).encode("utf-8"),
        file_name="qpcr_expression.csv",
        mime="text/csv"
    )
    st.download_button(
        texts["download_json"],
        data=merged_t.to_json(orient="records", force_ascii=False).encode("utf-8"),
        file_name="qpcr_expression.json",
        mime="application/json"
    )
    # --- Excel export ---
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        merged_t.to_excel(writer, sheet_name="Expression", index=False)
        # adjust column widths
        for idx, col in enumerate(merged_t.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            writer.book.active.column_dimensions[col_letter].width = max(15, len(col) + 2)
    output.seek(0)
    st.download_button(
        texts["download_excel"],
        data=output,
        file_name="qpcr_expression.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # QC summary
    qc = st.session_state.get("qc_summary", {})
    if qc:
        st.subheader(texts["qc_summary_header"])
        st.write(f"{texts['qc_total_wells']}: {qc['total_wells']}")
        st.write(f"{texts['qc_outliers']}: {qc['outliers']} ({qc['pass_rate']:.1f}%)")
        st.write(f"Outlier rate: {qc['outlier_rate']:.1f}%")

    # 📊 Generate chart based on selection
    if plot_type == texts["boxplot_header"]:
        st.subheader(texts["boxplot_header"])
        palette = px_colors.qualitative.Plotly
        tabs = st.tabs(merged_t["gene"].unique().tolist())
        for g, tb in zip(merged_t["gene"].unique(), tabs):
            with tb:
                sub = merged_t[merged_t["gene"] == g]
                fig_b = go.Figure()
                for i, cond in enumerate(sub["template"].unique()):
                    vals = sub[sub["template"] == cond]["expression"]
                    color = palette[i % len(palette)]
                    fig_b.add_trace(go.Box(
                        y=vals, name=cond,
                        fillcolor=color, line_color=color, marker_color=color,
                        boxpoints='all', jitter=0.3, pointpos=-1.8
                    ))
                fig_b.update_layout(title=f"Boxplot {g}", yaxis_title="expression", margin=dict(t=30,b=20))
                # Annotate pairwise significance for boxplot
                max_expr = sub["expression"].max()
                y_start = max_expr * 1.05
                h = max_expr * 0.03
                tpls = sub["template"].unique().tolist()
                for (a, b), p in st.session_state.get("pair_pvals", {}).get(g, []):
                    if p is None: continue
                    label = "*" if p < 0.05 else "ns"
                    # draw bracket
                    fig_b.add_shape(type="line",
                                    x0=a, x1=b, y0=y_start, y1=y_start,
                                    line=dict(color="red"))
                    fig_b.add_shape(type="line",
                                    x0=a, x1=a, y0=y_start, y1=y_start - h,
                                    line=dict(color="red"))
                    fig_b.add_shape(type="line",
                                    x0=b, x1=b, y0=y_start, y1=y_start - h,
                                    line=dict(color="red"))
                    # label at numeric midpoint
                    x0 = tpls.index(a)
                    x1 = tpls.index(b)
                    x_mid = (x0 + x1) / 2
                    fig_b.add_annotation(
                        x=x_mid, y=y_start + h * 1.2,
                        text=label, showarrow=False,
                        font=dict(size=16, color="red"),
                        xref="x", yref="y"
                    )
                    y_start += h * 2.5
                st.plotly_chart(fig_b, use_container_width=True)
                try:
                    img = pio.to_image(fig_b, format="png")
                    st.download_button(texts["download_boxplot"].format(g=g),
                                       data=img, file_name=f"boxplot_{g}.png", mime="image/png")
                except:
                    st.warning(texts["kaleido_warning"])
    else:
        st.subheader(texts["barplot_header"])
        tabs = st.tabs(merged_t["gene"].unique().tolist())
        for g, tb in zip(merged_t["gene"].unique(), tabs):
            with tb:
                sub = merged_t[merged_t["gene"] == g]
                stats = sub.groupby("template")["expression"].agg(["mean","std"]).reset_index()
                fig_bar = go.Figure(go.Bar(
                    x=stats["template"],
                    y=stats["mean"],
                    error_y=dict(type="data", array=stats["std"]),
                    marker_color="lightgrey"  # make bars grey instead of black
                ))
                # add scatter of raw values
                for i, cond in enumerate(stats["template"]):
                    y_vals = merged_t[(merged_t["gene"]==g) & (merged_t["template"]==cond)]["expression"]
                    fig_bar.add_trace(go.Scatter(
                        x=[cond]*len(y_vals),
                        y=y_vals,
                        mode='markers',
                        marker=dict(color='black', size=6),
                        showlegend=False
                    ))
                # Annotate pairwise significance for barplot
                max_val = stats["mean"].max()
                max_std = stats["std"].max()
                # Increase vertical offset so brackets sit higher above the error bars
                padding = max_std * 1.5
                y_start = max_val + padding
                # Use std-based height for consistent spacing
                h = max_std * 0.15
                tpls = stats["template"].tolist()
                for (a, b), p in st.session_state.get("pair_pvals", {}).get(g, []):
                    if p is None: continue
                    label = "*" if p < 0.05 else "ns"
                    # draw bracket
                    fig_bar.add_shape(type="line",
                                      x0=a, x1=b, y0=y_start, y1=y_start,
                                      line=dict(color="red"))
                    fig_bar.add_shape(type="line",
                                      x0=a, x1=a, y0=y_start, y1=y_start - h,
                                      line=dict(color="red"))
                    fig_bar.add_shape(type="line",
                                      x0=b, x1=b, y0=y_start, y1=y_start - h,
                                      line=dict(color="red"))
                    # label at numeric midpoint
                    x0 = tpls.index(a)
                    x1 = tpls.index(b)
                    x_mid = (x0 + x1) / 2
                    fig_bar.add_annotation(
                        x=x_mid, y=y_start + h * 1.8,
                        text=label, showarrow=False,
                        font=dict(size=16, color="red"),
                        xref="x", yref="y"
                    )
                    y_start += h * 3
                fig_bar.update_layout(title=f"Barplot {g}", yaxis_title="expression", margin=dict(t=30,b=20))
                st.plotly_chart(fig_bar, use_container_width=True)
                try:
                    img = pio.to_image(fig_bar, format="png")
                    st.download_button(
                        f"📥 Скачать barplot_{g}.png",
                        data=img,
                        file_name=f"barplot_{g}.png",
                        mime="image/png"
                    )
                except:
                    st.warning(texts["kaleido_warning"])


# -------------------- Full Report Generation --------------------
import tempfile
import matplotlib
# Ensure plotly can write images
import plotly.io as pio

st.sidebar.markdown("## 📑 Full Report")
if st.sidebar.button("Generate Full Report"):
    # Create PDF with FPDF
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", size=14)
    pdf.cell(0, 10, "qPCR Analysis Report", ln=True)

    # QC summary
    pdf.set_font("Arial", size=12)
    qc = st.session_state.get("qc_summary", {})
    pdf.cell(0, 8, f"Total wells: {qc.get('total_wells', 0)}", ln=True)
    pdf.cell(0, 8, f"Outliers: {qc.get('outliers', 0)} ({qc.get('outlier_rate', 0):.1f}%)", ln=True)
    pdf.ln(4)

    # Expression table
    merged_t = st.session_state.get("merged_t", pd.DataFrame())
    if "ΔCt" in merged_t.columns:
        merged_t = merged_t.rename(columns={"ΔCt": "dCt"})
    pdf.set_font("Courier", size=10)
    cols = merged_t.columns.tolist()
    header = "  ".join(c[:10].ljust(10) for c in cols)
    pdf.cell(0, 6, header, ln=True)
    for _, row in merged_t.iterrows():
        line = "  ".join(str(row[c])[:10].ljust(10) for c in cols)
        pdf.cell(0, 6, line, ln=True)
    pdf.ln(4)

    # Generate and embed both boxplot and barplot for each gene
    for g in merged_t["gene"].unique():
        sub = merged_t[merged_t["gene"] == g]
        # --- Generate Boxplot ---
        palette = px_colors.qualitative.Plotly
        fig_box = go.Figure()
        for i, cond in enumerate(sub["template"].unique()):
            vals = sub[sub["template"] == cond]["expression"]
            color = palette[i % len(palette)]
            fig_box.add_trace(go.Box(
                y=vals, name=cond, fillcolor=color,
                line_color=color, marker_color=color,
                boxpoints='all', jitter=0.3, pointpos=-1.8
            ))
        fig_box.update_layout(title=f"Boxplot {g}", yaxis_title="expression", margin=dict(t=30,b=20))
        # --- Generate Barplot ---
        stats_df = sub.groupby("template")["expression"].agg(["mean","std"]).reset_index()
        fig_bar = go.Figure(go.Bar(
            x=stats_df["template"], y=stats_df["mean"],
            error_y=dict(type="data", array=stats_df["std"]),
            marker_color="lightgrey"
        ))
        for cond in stats_df["template"]:
            vals = sub[sub["template"] == cond]["expression"]
            fig_bar.add_trace(go.Scatter(
                x=[cond]*len(vals), y=vals, mode="markers",
                marker=dict(color="black", size=6), showlegend=False
            ))
        fig_bar.update_layout(title=f"Barplot {g}", yaxis_title="expression", margin=dict(t=30,b=20))

        # Save boxplot to a temporary PNG file
        tmp_box = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        img_box_bytes = pio.to_image(fig_box, format="png", scale=2)
        tmp_box.write(img_box_bytes)
        tmp_box.flush()
        # Save barplot to a temporary PNG file
        tmp_bar = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
        img_bar_bytes = pio.to_image(fig_bar, format="png", scale=2)
        tmp_bar.write(img_bar_bytes)
        tmp_bar.flush()

        # Embed both images in PDF: boxplot then barplot
        pdf.ln(2)
        pdf.set_font("Arial", size=12)
        pdf.cell(0, 6, f"Boxplot for gene {g}", ln=True)
        pdf.image(tmp_box.name, w=100)
        pdf.ln(2)
        pdf.cell(0, 6, f"Barplot for gene {g}", ln=True)
        pdf.image(tmp_bar.name, w=100)
        pdf.ln(4)

        # Prepare HTML embed for both plots
        if "full_html_parts" not in locals():
            full_html_parts = ['<h1>qPCR Analysis Report</h1>']
            full_html_parts.append("<h2>QC Summary</h2><ul>"
                                   f"<li>Total wells: {qc.get('total_wells',0)}</li>"
                                   f"<li>Outliers: {qc.get('outliers',0)} ({qc.get('outlier_rate',0):.1f}%)</li>"
                                   "</ul>")
            full_html_parts.append("<h2>Expression Results</h2>")
            full_html_parts.append(merged_t.to_html(index=False, classes='table table-striped'))
        # Boxplot HTML
        img_box_b64 = base64.b64encode(img_box_bytes).decode("utf-8")
        full_html_parts.append(f"<h2>Boxplot for gene {g}</h2>")
        full_html_parts.append(f'<img src="data:image/png;base64,{img_box_b64}" style="max-width:600px;"/>')
        # Barplot HTML
        img_bar_b64 = base64.b64encode(img_bar_bytes).decode("utf-8")
        full_html_parts.append(f"<h2>Barplot for gene {g}</h2>")
        full_html_parts.append(f'<img src="data:image/png;base64,{img_bar_b64}" style="max-width:600px;"/>')

    # Finalize PDF
    pdf_bytes = pdf.output(dest='S').encode('latin-1')
    buf_pdf = io.BytesIO(pdf_bytes)
    buf_pdf.seek(0)
    st.sidebar.download_button(
        "Download report (PDF)",
        data=buf_pdf,
        file_name="qpcr_full_report.pdf",
        mime="application/pdf"
    )
    # Upload PDF report to Google Drive
    pdf_drive_file = drive.CreateFile({
        "title": "qpcr_full_report.pdf",
        "parents": [{"id": FOLDER_ID}],
        "mimeType": "application/pdf"
    })
    pdf_drive_file.ContentBinary = buf_pdf.getvalue()
    pdf_drive_file.Upload()
    st.sidebar.success("PDF report uploaded to Google Drive")

    # Finalize HTML report
    full_html = "<html><head><title>qPCR Report</title></head><body>" + "".join(full_html_parts) + "</body></html>"
    st.sidebar.download_button(
        "Download report (HTML)",
        data=full_html,
        file_name="qpcr_full_report.html",
        mime="text/html"
    )
    # Upload HTML report to Google Drive
    html_drive_file = drive.CreateFile({
        "title": "qpcr_full_report.html",
        "parents": [{"id": FOLDER_ID}],
        "mimeType": "text/html"
    })
    html_drive_file.SetContentString(full_html)
    html_drive_file.Upload()
    st.sidebar.success("HTML report uploaded to Google Drive")